#!/usr/bin/env python3
import re, argparse, os, sys, math, datetime, json
from typing import Dict, Any, List, Tuple, Optional
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    import requests
except Exception:
    requests = None  # if requests not installed, we’ll just skip adverse-news

# ---------- Helpers ----------
def read_pdf_text(path: str) -> str:
    with pdfplumber.open(path) as pdf:
        return "\n".join([(page.extract_text() or "") for page in pdf.pages])

HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
SECTION_FILL = PatternFill("solid", fgColor="E8F5E9")
BORDER_THIN = Border(left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"))

TODAY = datetime.date.today()

def parse_date(d: Optional[str]) -> Optional[datetime.date]:
    if not d: return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.datetime.strptime(d.strip(), fmt).date()
        except Exception:
            continue
    return None

def years_between(d1: datetime.date, d2: datetime.date) -> float:
    return (d2 - d1).days / 365.25

def set_col_widths(ws, widths: List[Tuple[int, float]]):
    for col, width in widths:
        ws.column_dimensions[get_column_letter(col)].width = width

def add_section(ws, title: str, kv: List[Tuple[str, Any]], start_row: int, cols: int = 6) -> int:
    kv_nonempty = [(k,v) for k,v in kv if v not in (None, "", [])]
    if not kv_nonempty: return start_row
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=cols)
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = Font(bold=True, size=12); cell.alignment = Alignment(horizontal="left", vertical="center"); cell.fill = SECTION_FILL
    start_row += 1
    ws.cell(row=start_row, column=1, value="Field").font = Font(bold=True)
    ws.cell(row=start_row, column=2, value="Value").font = Font(bold=True)
    ws.cell(row=start_row, column=1).fill = HEADER_FILL; ws.cell(row=start_row, column=2).fill = HEADER_FILL
    ws.cell(row=start_row, column=1).border = BORDER_THIN; ws.cell(row=start_row, column=2).border = BORDER_THIN
    start_row += 1
    for key, val in kv_nonempty:
        ws.cell(row=start_row, column=1, value=key).border = BORDER_THIN
        if isinstance(val, list):
            lines = []
            for x in val:
                if isinstance(x, dict): lines.append("- " + ", ".join(f"{kk}: {vv}" for kk, vv in x.items()))
                else: lines.append(f"- {x}")
            val = "\n".join(lines)
        ws.cell(row=start_row, column=2, value=val).border = BORDER_THIN
        ws.cell(row=start_row, column=2).alignment = Alignment(wrap_text=True)
        start_row += 1
    start_row += 1
    return start_row

def add_table(ws, title: str, columns: List[str], rows: List[List[Any]], start_row: int, cols: int = 6) -> int:
    if not rows: return start_row
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=cols)
    cell = ws.cell(row=start_row, column=1, value=title); cell.font = Font(bold=True, size=12); cell.fill = SECTION_FILL
    start_row += 1
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=c, value=name); cell.font = Font(bold=True); cell.fill = HEADER_FILL; cell.border = BORDER_THIN
    start_row += 1
    for r in rows:
        for c, val in enumerate(r, start=1):
            cell = ws.cell(row=start_row, column=c, value=val); cell.border = BORDER_THIN; cell.alignment = Alignment(wrap_text=True)
        start_row += 1
    start_row += 1
    return start_row

# Slice text from a header until the next known header
ALLCAPS_LINE = re.compile(r'^[A-Z][A-Z0-9 ()/&\-\':]{2,}$')
KNOWN_HEADERS_UP = [
    "\nLITIGATION - AS PLAINTIFF",
    "\nLITIGATION - AS DEFENDANT",
    "\nLITIGATION",
    "\nBANKRUPTCY",
    "\nBANKRUPTCY / WINDING UP",
    "\nBANKRUPTCY & WINDING UP",
]
def slice_block(text: str, header_upper: str) -> str:
    up = text.upper()
    i = up.find(header_upper)
    if i == -1:
        return ""
    start = i + len(header_upper)
    nexts = [up.find(h, start) for h in KNOWN_HEADERS_UP]
    nexts = [x for x in nexts if x != -1]
    end = min(nexts) if nexts else len(up)
    block = text[start:end]
    blk_lines = block.splitlines()
    if blk_lines and ALLCAPS_LINE.fullmatch(blk_lines[0].strip()):
        blk_lines = blk_lines[1:]
    return "\n".join(blk_lines).strip()

# ---------- PARSERS ----------
def parse_stars(text: str) -> Dict[str, Any]:
    def grab(pattern, flags=0):
        m = re.search(pattern, text, flags); return m.group(1).strip() if m else None
    data: Dict[str, Any] = {}
    data['Lot Number'] = grab(r'Lot Number\s*:\s*(.+)')
    block = re.search(r'Property Address\s*:\s*(.*?)\n\s*\n', text, re.S)
    if block: data['Property Address'] = " ".join([ln.strip() for ln in block.group(1).splitlines()])
    lot_area = grab(r'Lot Area\s*:\s*([0-9]+(?:\.[0-9]+)?)\s*SqM')
    if lot_area:
        try:
            sqm = float(lot_area); data['Lot Area (SqM)'] = sqm; data['Lot Area (SqFt)'] = round(sqm * 10.7639)
        except: data['Lot Area (SqM)'] = lot_area
    data['State Title Tenure'] = grab(r'State Title Tenure\s*:\s*(.+)')
    data['Lease Duration'] = grab(r'Lease Duration\s*:\s*(.+)')
    data['Commencement Date'] = grab(r'Commencement Date\s*:\s*([0-9]{2}/[0-9]{2}/[0-9]{4})')
    data['Expiry Date'] = grab(r'State Title Expiry Date\s*:\s*([0-9]{2}/[0-9]{2}/[0-9]{4}|[0-9]{2}/[0-9]{4})')

    owners = re.findall(r'Name\s*:\s*([A-Z0-9 ()\/\.-]+)\n\s*Address', text)
    if owners: data['Owners'] = ", ".join([o.strip() for o in owners])

    m = re.search(r'\b(EXECUTIVE CONDOMINIUM|APARTMENT|HDB|LANDED|CONDOMINIUM)\b', text)
    if m: data['Property Type'] = m.group(1).strip()
    m = re.search(r'\b(\d{6})\b', data.get('Property Address',''))
    if m: data['Postal Code'] = m.group(1)

    # Encumbrances
    encs: List[Dict[str, Any]] = []
    for m in re.finditer(r'APPLICATION TO NOTIFY CHARGE\s+([A-Z0-9/]+)\s+lodge[d]?\s+on\s+([0-9/]+)\s+at\s+([0-9:]+)(.*?)(?=\n\d+\s|MORTGAGE\b|$)', text, re.S|re.I):
        rec = {"Type": "Application to Notify Charge", "Instrument No": m.group(1), "Lodged On": m.group(2), "Lodged Time": m.group(3)}
        m2 = re.search(r'CHARGEE\s*-+\s*\n\s*(.+)', m.group(4))
        if m2: rec["Counterparty"] = m2.group(1).strip()
        m2 = re.search(r'Type of Charge\s*:\s*(.+)', m.group(4))
        if m2: rec["Charge Type"] = m2.group(1).strip()
        m2 = re.search(r'NOTIFIED ON\s*:\s*([0-9/]+)', m.group(4))
        if m2: rec["Registered/Notified On"] = m2.group(1).strip()
        encs.append(rec)
    for m in re.finditer(r'MORTGAGE\s+([A-Z0-9/]+)\s+lodge[d]?\s+on\s+([0-9/]+)\s+at\s+([0-9:]+)(.*?)(?=\n\d+\s|APPLICATION TO NOTIFY CHARGE\b|$)', text, re.S|re.I):
        rec = {"Type": "Mortgage", "Instrument No": m.group(1), "Lodged On": m.group(2), "Lodged Time": m.group(3)}
        m2 = re.search(r'MORTGAGEE\s*-+\s*\n\s*(.+)', m.group(4))
        if m2: rec["Counterparty"] = m2.group(1).strip()
        m2 = re.search(r'REGISTERED ON\s*:\s*([0-9/]+)', m.group(4))
        if m2: rec["Registered/Notified On"] = m.group(1).strip() if m else None
        encs.append(rec)
    if encs: data['Encumbrances'] = encs
    return data

# ---- SCCB (ACRA) ----
CASE_TOKEN = re.compile(r'\b(?:HC|DC|MC|OS|SUM|SIC|AR|AD|MAG)/[0-9A-Z]+/\d{2,4}\b', re.I)
NOREC_WORDS = ("NO RECORD FOUND", "NO RECORDS FOUND", "NO RECORD", "NO RECORDS", "NONE", "NIL")
HEADER_TOKENS = set([
    "CASE","CASE NO","CASENO","CASE NUMBER","NO","NUMBER","COURT","CITATION","DATE","FILED","FILING DATE",
    "HEARING DATE","PARTY","PARTIES","PLAINTIFF","DEFENDANT","RESPONDENT","STATUS","OUTCOME","REMARKS",
    "REFERENCE","AMOUNT","SUMS"
])

def looks_like_only_headers(block: str) -> bool:
    if not block or not block.strip(): return True
    for raw in block.splitlines():
        ln = raw.strip()
        if not ln: continue
        up = re.sub(r'[^A-Z ]+', ' ', ln.upper()).strip()
        words = [w for w in up.split() if w]
        if words and all((w in HEADER_TOKENS) or (len(w) <= 2) for w in words):
            continue
        if re.fullmatch(r'[-_ ]{2,}', ln): continue
        if re.fullmatch(r'(N/?A|NA)', ln, flags=re.I): continue
        return False
    return True

def decide_lit_status_side(block: str) -> str:
    norm = re.sub(r"\s+", " ", (block or "")).upper()
    if not norm or any(kw in norm for kw in NOREC_WORDS):
        return "NO RECORD FOUND"
    if looks_like_only_headers(block):
        return "NO RECORD FOUND"
    if CASE_TOKEN.search(norm):
        return "Present"
    if re.search(r'\b\d{1,4}/\d{2,4}\b', norm):  # numeric row fallback
        return "Present"
    return "NO RECORD FOUND"

def parse_sccb(text: str) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    m = re.search(r'REQUESTED INDIVIDUAL NAME\s*:\s*(.+)', text)
    if m: data['Individual Name'] = m.group(1).strip()
    if 'Individual Name' not in data:
        m = re.search(r'INDIVIDUAL NAME\s*:\s*(.+)', text)
        if m: data['Individual Name'] = m.group(1).strip()
    m = re.search(r'NRIC\s*/\s*ID\s*:\s*([A-Z0-9]+)', text)
    if m: data['NRIC'] = m.group(1).strip()
    m = re.search(r'ADDRESS\s*CHANGED DATE.*?\n(\d{2}/\d{2}/\d{4})\s+(.+)', text)
    if m: 
        data['ACRA Address'] = m.group(2).strip(); data['Address Updated'] = m.group(1)
    m = re.search(r'CURRENT COMPANIES.*?\n([0-9A-Z]+)\s+(.+?)\n.*?\n(\d{2}/\d{2}/\d{4}).*?(DIRECTOR|MANAGER|PARTNER|SHAREHOLDER)', text, re.S)
    if m:
        data['Current Company UEN'] = m.group(1).strip()
        data['Current Company'] = m.group(2).strip()
        data['Appointment Date'] = m.group(3).strip()
        data['Position'] = m.group(4).strip()

    p_block = slice_block(text, "LITIGATION - AS PLAINTIFF")
    d_block = slice_block(text, "LITIGATION - AS DEFENDANT")
    if not p_block and not d_block:
        generic = slice_block(text, "LITIGATION")
        upg = generic.upper()
        if "AS PLAINTIFF" in upg or "AS PLAINTIFF/CLAIMANT" in upg:
            p_block = slice_block("LITIGATION - AS PLAINTIFF\n" + generic, "LITIGATION - AS PLAINTIFF")
        if "AS DEFENDANT" in upg:
            d_block = slice_block("LITIGATION - AS DEFENDANT\n" + generic, "LITIGATION - AS DEFENDANT")
        if not p_block and not d_block:
            p_block = generic; d_block = generic

    p_status = decide_lit_status_side(p_block)
    d_status = decide_lit_status_side(d_block)

    if p_status == "NO RECORD FOUND" and d_status == "NO RECORD FOUND":
        data['SCCB Litigation Status'] = "NO RECORD FOUND"; data['SCCB Litigation Sides'] = "None"
    elif p_status != "NO RECORD FOUND" and d_status != "NO RECORD FOUND":
        data['SCCB Litigation Status'] = "Present"; data['SCCB Litigation Sides'] = "Both"
    elif p_status != "NO RECORD FOUND":
        data['SCCB Litigation Status'] = "Present"; data['SCCB Litigation Sides'] = "Plaintiff"
    else:
        data['SCCB Litigation Status'] = "Present"; data['SCCB Litigation Sides'] = "Defendant"

    data['_SCCB_LIT_RAW_P'] = p_block if data['SCCB Litigation Status'] == "Present" else ""
    data['_SCCB_LIT_RAW_D'] = d_block if data['SCCB Litigation Status'] == "Present" else ""

    b_block = slice_block(text, "BANKRUPTCY")
    bnorm = re.sub(r"\s+", " ", (b_block or "")).upper()
    data['SCCB Bankruptcy Status'] = "NIL" if (not bnorm or any(w in bnorm for w in ("NIL","NO RECORD","NO RECORDS","NONE"))) else "Present"
    data['_SCCB_BKY_RAW'] = b_block if data['SCCB Bankruptcy Status'] == "Present" else ""
    return data

# ---- CBS (Consumer Credit) ----
CBS_HEADER_WORDS = {"NARRATIVES", "NARRATIVE", "ACCOUNT NARRATIVES", "ACCOUNTS NARRATIVE", "DATE LOADED TYPE"}
def clean_cbs_narratives(block: str, max_lines: int = 5) -> Optional[str]:
    if not block: return None
    out, seen = [], set()
    for ln in (ln.strip() for ln in block.splitlines()):
        if not ln: continue
        up = ln.upper()
        if up in CBS_HEADER_WORDS: continue
        if re.fullmatch(r'[-_ ]{2,}', ln): continue
        if re.fullmatch(r'[A-Za-z]', ln): continue  # drop single-letter strays (e.g., "A")
        ln = re.sub(r'\s+', ' ', ln)
        if ln in seen: continue
        seen.add(ln); out.append(ln)
        if len(out) >= max_lines: break
    return "\n".join("• " + ln for ln in out) if out else None

def parse_cbs(text: str) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    m = re.search(r'Name:\s*(.+?)\s{2,}Date of Earliest', text, re.S)
    if m: data['CBS Name'] = " ".join(m.group(1).split())
    m = re.search(r'ID Type:\s*(\w+)\s{2,}', text, re.S)
    if m: data['ID Type'] = m.group(1).strip()
    m = re.search(r'ID Number:\s*([A-Z0-9]+)\s{2,}', text, re.S)
    if m: data['CBS NRIC/ID'] = m.group(1).strip()
    m = re.search(r'Date of Birth:\s*([0-9/]+)', text);  data['Date of Birth'] = m.group(1) if m else data.get('Date of Birth')
    m = re.search(r'Postal Code:\s*([0-9]{6})', text);   data['CBS Postal Code'] = m.group(1) if m else data.get('CBS Postal Code')
    m = re.search(r'Score[.\s:]*([0-9]{3,4})', text, re.I); data['Credit Score'] = m.group(1) if m else data.get('Credit Score')
    m = re.search(r'Risk\s*Grade[.\s:]*([A-Z]{1,2}[0-9]?)', text, re.I); data['Risk Grade'] = m.group(1) if m else data.get('Risk Grade')
    m = re.search(r'Total\s*Credit\s*Limit\s*[:\s]\s*\$?\s*([0-9,]+\.[0-9]{2}|[0-9,]+)', text, re.I); data['Total Credit Limit'] = m.group(1) if m else data.get('Total Credit Limit')
    m = re.search(r'Total\s*Outstanding\s*Balance\s*[:\s]\s*\$?\s*([0-9,]+\.[0-9]{2}|[0-9,]+)', text, re.I); data['Total Outstanding Balance'] = m.group(1) if m else data.get('Total Outstanding Balance')
    m = re.search(r'Previous\s*Enquiries.*?Last\s*12\s*Months\s*[:\s]\s*([0-9]+)', text, re.I|re.S); data['Previous Enquiries (12m)'] = m.group(1) if m else data.get('Previous Enquiries (12m)')

    for hp in ["NARRATIVES", "NARRATIVE", "ACCOUNT NARRATIVES"]:
        blk = slice_block(text, hp)
        if blk and blk.strip():
            data['CBS Narratives'] = clean_cbs_narratives(blk)
            break

    m = re.search(r'Default Records.*?(\d{2}/\d{2}/\d{4})\s+([0-9,]+\.[0-9]{2})\s+([0-9,]+\.[0-9]{2})', text, re.S)
    if m:
        data['Default Loaded'] = m.group(1); data['Default Original Amount'] = m.group(2); data['Default Balance'] = m.group(3)

    m = re.search(r'Bankruptcy Number.*?\n([0-9]+)\s+([0-9/]+).*?\n.*?([0-9]+)\s+([0-9/]+)', text, re.S)
    if m:
        data['Bankruptcy Order No'] = m.group(1); data['Bankruptcy Order Date'] = m.group(2)
        data['Bankruptcy Discharge No'] = m.group(3); data['Bankruptcy Discharge Date'] = m.group(4)
    return data

# ---------- Workbook ----------
def summarize_encumbrances(encs: List[Dict[str,Any]]) -> Optional[str]:
    """Compact, tidy multi-line summary for the Summary sheet with dedupe and ≤3 bullets."""
    if not encs: return None
    seen = set(); uniq: List[Dict[str,Any]] = []
    for rec in encs:
        key = (
            (rec.get("Type") or "").strip().lower(),
            (rec.get("Instrument No") or "").strip().lower(),
            (rec.get("Counterparty") or "").strip().lower(),
            (rec.get("Lodged On") or "").strip().lower(),
            (rec.get("Registered/Notified On") or "").strip().lower(),
        )
        if key in seen: continue
        seen.add(key); uniq.append(rec)
    bullets = []
    for rec in uniq[:3]:
        t = (rec.get("Type") or "").strip().title()
        cp = (rec.get("Counterparty") or "").strip()
        instr = (rec.get("Instrument No") or "").strip()
        lodg = (rec.get("Lodged On") or "").strip()
        notif = (rec.get("Registered/Notified On") or "").strip()
        parts = []
        if t: parts.append(t)
        if cp: parts.append(f"({cp})")
        if instr: parts.append(f"#{instr}")
        if lodg: parts.append(f"lodged {lodg}")
        if notif: parts.append(f"registered/notified {notif}")
        bullet = "• " + " ".join(parts).strip()
        if bullet.strip("• ").strip():
            bullets.append(bullet)
    total = len(uniq)
    if total <= 3: return "\n".join(bullets) if bullets else None
    more = total - 3
    if more > 0: bullets.append(f"• +{more} more item(s)")
    return "\n".join(bullets) if bullets else None

# ---------- Adverse-news (optional) ----------
# --- replace your NEG_TERMS/NEWSY_ALLOWLIST/helpers + adverse_news() with this ---

import urllib.parse

NEG_TERMS = [
    "bankrupt", "bankruptcy", "winding up", "insolvency",
    "fraud", "scam", "cheat", "embezzle", "forgery",
    "lawsuit", "sued", "prosecuted", "charged", "convicted",
    "court", "litigation", "police probe", "investigation",
    "criminal breach of trust", "cbt"
]

NEWSY_ALLOWLIST = {
    "straitstimes.com", "todayonline.com", "channelnewsasia.com", "businesstimes.com.sg",
    "asiaone.com", "reuters.com", "bloomberg.com", "bbc.com", "scmp.com",
}

# hard excludes (tune this list as you see false positives)
HARD_EXCLUDE = {
    "instagram.com", "facebook.com", "linkedin.com", "x.com", "twitter.com",
    "isca.org.sg", "youtube.com", "tiktok.com", "medium.com", "wikipedia.org",
}

def _domain(url: str) -> str:
    try:
        netloc = urllib.parse.urlparse(url).netloc.lower()
        # normalize to last 2-3 labels
        parts = netloc.split(".")
        if len(parts) >= 3 and parts[-2] in {"com","co","org","net","gov","edu","sg","my"}:
            return ".".join(parts[-3:])
        return ".".join(parts[-2:])
    except Exception:
        return ""

def _contains_name(text: str, name: str) -> bool:
    tokens = [t for t in re.split(r"\s+", name.strip()) if len(t) >= 2]
    if not tokens: return False
    up = text.lower()
    # require first two tokens somewhere
    need = tokens[:2] if len(tokens) >= 2 else tokens
    return all(tok.lower() in up for tok in need)

def _has_neg_term(text: str) -> bool:
    up = text.lower()
    return any(term in up for term in NEG_TERMS)

def _name_neg_near(text: str, name: str, window: int = 80) -> bool:
    """Require any neg-term to appear within `window` chars of the name occurrence."""
    t = text.lower()
    name_l = name.lower()
    # allow slight punctuation/space differences in name
    name_re = re.sub(r"\s+", r"\\s+", re.escape(name_l))
    m = re.search(name_re, t)
    if not m:
        return False
    start, end = m.start(), m.end()
    left = max(0, start - window)
    right = min(len(t), end + window)
    neighborhood = t[left:right]
    return _has_neg_term(neighborhood)

def adverse_news(name: str, limit: int = 5) -> List[Dict[str, str]]:
    api_key = os.environ.get("GOOGLE_CSE_API_KEY")
    cx      = os.environ.get("GOOGLE_CSE_ENGINE_ID")
    if not (api_key and cx and name and requests):
        return []

    # build query: exact name + (neg terms), try to avoid PDFs and socials
    neg_block = " OR ".join([f'"{t}"' if " " in t else t for t in NEG_TERMS])
    q = f'"{name}" ({neg_block}) -filetype:pdf -site:linkedin.com -site:instagram.com -site:facebook.com'

    try:
        r = requests.get(
            "https://www.googleapis.com/customsearch/v1",
            params={
                "key": api_key,
                "cx": cx,
                "q": q,
                "num": min(limit * 5, 10),  # fetch more, we'll filter down
                "safe": "active",
                "gl": "sg",
                "lr": "lang_en",
                "dateRestrict": "y1",
            },
            timeout=10
        )
        data = r.json() if r.ok else {}
        raw_items = data.get("items", [])
    except Exception:
        return []

    filtered = []
    for it in raw_items:
        title = (it.get("title") or "").strip()
        snippet = (it.get("snippet") or "").strip()
        link = (it.get("link") or "").strip()
        if not (title and link):
            continue

        dom = _domain(link)
        if dom in HARD_EXCLUDE:
            continue

        combo = f"{title} {snippet}"

        # must contain name and a negative term
        if not _contains_name(combo, name):
            continue
        if not _has_neg_term(combo):
            continue

        # proximity test: neg term near the name (reduce bios/benign mentions)
        if not _name_neg_near(combo, name, window=80):
            # if domain is clearly newsy, allow even if not near, else skip
            if dom not in NEWSY_ALLOWLIST:
                continue

        score = (2 if dom in NEWSY_ALLOWLIST else 0)
        filtered.append({"title": title, "snippet": snippet, "link": link, "_score": score})

    filtered.sort(key=lambda x: x["_score"], reverse=True)
    return [{k: v for k, v in x.items() if not k.startswith("_")} for x in filtered[:limit]]


# ---------- Build Workbook ----------
def build_workbook(stars: Dict[str,Any], sccb: Dict[str,Any], cbs: Dict[str,Any], out_path: str, attachments: List[str], adverse_hits: Optional[List[Dict[str,str]]] = None) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "Case Data"
    set_col_widths(ws, [(1, 34), (2, 70), (3, 14), (4, 14), (5, 14), (6, 14)])
    row = 1
    dob = parse_date(cbs.get("Date of Birth")); age = math.floor(years_between(dob, TODAY)) if dob else None
    lease_expiry = parse_date(stars.get("Expiry Date")); yrs_rem = round(years_between(TODAY, lease_expiry), 1) if lease_expiry else None

    # Property Details
    prop_fields = [
        ("Property Address", stars.get("Property Address")),
        ("Postal Code", stars.get("Postal Code")),
        ("Property Type", stars.get("Property Type")),
        ("Tenure", stars.get("State Title Tenure")),
        ("Leasehold Years", stars.get("Lease Duration")),
        ("Lease Start", stars.get("Commencement Date")),
        ("Lease Expiry", stars.get("Expiry Date")),
        ("Years Remaining (approx.)", yrs_rem),
        ("Lot Area (SqM)", stars.get("Lot Area (SqM)")),
        ("Lot Area (SqFt)", stars.get("Lot Area (SqFt)")),
        ("Lot Number", stars.get("Lot Number")),
        ("Owners", stars.get("Owners")),
    ]
    row = add_section(ws, "Property Details", prop_fields, row, cols=6)

    # Encumbrances (full table)
    enc_rows = []
    for rec in stars.get("Encumbrances", []) or []:
        enc_rows.append([
            rec.get("Type",""),
            rec.get("Instrument No",""),
            rec.get("Counterparty",""),
            rec.get("Charge Type",""),
            rec.get("Lodged On",""),
            rec.get("Registered/Notified On","")
        ])
    row = add_table(ws, "Encumbrances: Charges & Mortgages",
                    ["Type","Instrument No","Counterparty","Charge Type","Lodged On","Registered/Notified On"],
                    enc_rows, row, cols=6)

    # Loan Detail & Outstanding
    loan_fields = [
        ("Credit Score", cbs.get("Credit Score")),
        ("Risk Grade", cbs.get("Risk Grade")),
        ("Total Credit Limit", cbs.get("Total Credit Limit")),
        ("Total Outstanding Balance", cbs.get("Total Outstanding Balance")),
        ("Previous Enquiries (Last 12m)", cbs.get("Previous Enquiries (12m)")),
        ("Default Record Loaded", cbs.get("Default Loaded")),
        ("Default Original Amount", cbs.get("Default Original Amount")),
        ("Default Balance", cbs.get("Default Balance")),
        ("Bankruptcy Order No", cbs.get("Bankruptcy Order No")),
        ("Bankruptcy Order Date", cbs.get("Bankruptcy Order Date")),
        ("Bankruptcy Discharge No", cbs.get("Bankruptcy Discharge No")),
        ("Bankruptcy Discharge Date", cbs.get("Bankruptcy Discharge Date")),
        ("CBS Narratives", cbs.get("CBS Narratives")),
    ]
    row = add_section(ws, "Loan Detail & Outstanding", loan_fields, row, cols=6)

    # Individual Borrower / PG
    borrower_fields = [
        ("Name (CBS)", cbs.get("CBS Name")),
        ("NRIC/ID (CBS)", cbs.get("CBS NRIC/ID")),
        ("Date of Birth", cbs.get("Date of Birth")),
        ("Age (years)", age),
        ("Postal Code (CBS)", cbs.get("CBS Postal Code")),
        ("Name (ACRA)", sccb.get("Individual Name")),
        ("NRIC/ID (ACRA)", sccb.get("NRIC")),
        ("Residential / ACRA Address", sccb.get("ACRA Address")),
        ("Address Updated", sccb.get("Address Updated")),
        ("Current Company", sccb.get("Current Company")),
        ("UEN", sccb.get("Current Company UEN")),
        ("Position", sccb.get("Position")),
        ("Appointment Date", sccb.get("Appointment Date")),
        ("SCCB Bankruptcy Status", sccb.get("SCCB Bankruptcy Status")),
        ("SCCB Litigation Status", sccb.get("SCCB Litigation Status")),
        ("SCCB Litigation Sides", sccb.get("SCCB Litigation Sides")),
    ]
    row = add_section(ws, "Individual Borrower / PG", borrower_fields, row, cols=6)

    # Loan Summary
    risk_flags = []
    if cbs.get("Default Balance"): risk_flags.append("Default record present")
    if cbs.get("Bankruptcy Order No"): risk_flags.append("Bankruptcy history" + (" (discharged)" if cbs.get("Bankruptcy Discharge Date") else ""))
    if cbs.get("CBS Narratives"): risk_flags.append("Narratives present in CBS")
    if sccb.get("SCCB Litigation Status") == "Present": risk_flags.append(f"SCCB Litigation present ({sccb.get('SCCB Litigation Sides')})")
    if sccb.get("SCCB Bankruptcy Status") == "Present": risk_flags.append("SCCB Bankruptcy present")

    enc_compact = summarize_encumbrances(stars.get("Encumbrances", []) or [])
    summary_fields = [
        ("Overall Risk Flags", ", ".join(risk_flags) if risk_flags else "None seen in docs"),
        ("Borrower Name", cbs.get("CBS Name") or sccb.get("Individual Name")),
        ("NRIC/ID", cbs.get("CBS NRIC/ID") or sccb.get("NRIC")),
        ("Property Address", stars.get("Property Address")),
        ("Encumbrances (compact)", enc_compact),
        ("SCCB Litigation", sccb.get("SCCB Litigation Status") + (f" ({sccb.get('SCCB Litigation Sides')})" if sccb.get("SCCB Litigation Status") == "Present" else "")),
        ("SCCB Bankruptcy", sccb.get("SCCB Bankruptcy Status")),
    ]
    # Adverse news one-liner
    if adverse_hits:
        summary_fields.append(("Adverse News", f"{len(adverse_hits)} item(s)"))  # compact line

    row = add_section(ws, "Loan Summary", summary_fields, row, cols=6)

    # Optional: Adverse News Section in Case Data
    if adverse_hits:
        bullets = []
        for h in adverse_hits[:5]:
            title = (h.get("title") or "").strip()
            link  = (h.get("link") or "").strip()
            snip  = (h.get("snippet") or "").strip()
            if title and link:
                bullets.append(f"{title} — {link}\n{snip}")
        row = add_section(ws, "Adverse News (Top Hits)", [("Items", "\n\n".join("• " + b for b in bullets))], row, cols=6)

    # Broker Info, Attachments, Revenue placeholders
    broker_fields = [("Broker Name", None), ("Contact", None), ("Notes", None)]
    row = add_section(ws, "Broker Information", broker_fields, row, cols=6)
    attach_fields = [("Attachment", a) for a in attachments]
    row = add_section(ws, "Attachments", attach_fields, row, cols=6)
    revenue_fields = [("Property Valuation", None), ("Brokerage Fee (%)", None), ("Brokerage Fee ($)", None), ("Other Fees", None)]
    row = add_section(ws, "Revenue", revenue_fields, row, cols=6)

    ws.freeze_panes = "A2"

    # Summary sheet (one-pager)
    ws2 = wb.create_sheet("Summary"); set_col_widths(ws2, [(1, 36), (2, 70)]); r = 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell = ws2.cell(row=r, column=1, value="Case Summary"); cell.font = Font(bold=True, size=14); cell.alignment = Alignment(horizontal="left"); cell.fill = SECTION_FILL
    r += 2
    def add_kv(row_label, value):
        nonlocal r
        if value in (None, "", []): return
        ws2.cell(row=r, column=1, value=row_label).font = Font(bold=True); ws2.cell(row=r, column=1).fill = HEADER_FILL; ws2.cell(row=r, column=1).border = BORDER_THIN
        ws2.cell(row=r, column=2, value=value).border = BORDER_THIN; ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True); r += 1

    add_kv("Borrower", (cbs.get("CBS Name") or sccb.get("Individual Name") or ""))
    add_kv("NRIC/ID", (cbs.get("CBS NRIC/ID") or sccb.get("NRIC") or ""))
    add_kv("DOB / Age", f"{cbs.get('Date of Birth','')} / {math.floor(years_between(parse_date(cbs.get('Date of Birth')), TODAY)) if parse_date(cbs.get('Date of Birth')) else ''}")
    add_kv("Property Address", stars.get("Property Address"))
    if stars.get('Lot Number') or stars.get('Lot Area (SqM)'): add_kv("Lot / Area (SqM)", f"{stars.get('Lot Number','')} / {stars.get('Lot Area (SqM)','')}")
    add_kv("Area (SqFt)", stars.get('Lot Area (SqFt)'))
    add_kv("Tenure / Lease", f"{stars.get('State Title Tenure','')} | {stars.get('Lease Duration','')}")
    add_kv("Lease Commence → Expiry", f"{stars.get('Commencement Date','')} → {stars.get('Expiry Date','')}")
    add_kv("Years Remaining (approx.)", yrs_rem)
    add_kv("Encumbrances", enc_compact)
    add_kv("Credit Score / Grade", " / ".join([x for x in [cbs.get('Credit Score'), cbs.get('Risk Grade')] if x]))
    add_kv("Totals (Limit / O/S)", " / ".join([x for x in [cbs.get('Total Credit Limit'), cbs.get('Total Outstanding Balance')] if x]))
    add_kv("Defaults", f"Loaded {cbs.get('Default Loaded','')} | Balance {cbs.get('Default Balance','')}" if cbs.get("Default Balance") else "None seen")
    add_kv("SCCB Litigation", sccb.get("SCCB Litigation Status") + (f" ({sccb.get('SCCB Litigation Sides')})" if sccb.get("SCCB Litigation Status") == "Present" else ""))
    add_kv("SCCB Bankruptcy", sccb.get("SCCB Bankruptcy Status"))
    if adverse_hits: add_kv("Adverse News", f"{len(adverse_hits)} item(s)")
    add_kv("Attachments", ", ".join(attachments))

    # Hidden raw SCCB sheet (only if records exist)
    if sccb.get("SCCB Litigation Status") == "Present" or sccb.get("SCCB Bankruptcy Status") == "Present":
        ws3 = wb.create_sheet("SCCB_Raw")
        set_col_widths(ws3, [(1, 30), (2, 100)])
        rr = 1
        def add_block(title, content):
            nonlocal rr
            if content:
                ws3.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
                c = ws3.cell(row=rr, column=1, value=title); c.font=Font(bold=True); c.fill=SECTION_FILL
                rr += 1
                ws3.cell(row=rr, column=1, value="Text").font=Font(bold=True); ws3.cell(row=rr, column=1).fill=HEADER_FILL
                ws3.cell(row=rr, column=2, value="Content").font=Font(bold=True); ws3.cell(row=rr, column=2).fill=HEADER_FILL
                rr += 1
                ws3.cell(row=rr, column=1, value="Raw").border=BORDER_THIN
                ws3.cell(row=rr, column=2, value=content).border=BORDER_THIN; ws3.cell(row=rr, column=2).alignment=Alignment(wrap_text=True)
                rr += 2
        add_block("Litigation — As Plaintiff", sccb.get("_SCCB_LIT_RAW_P"))
        add_block("Litigation — As Defendant", sccb.get("_SCCB_LIT_RAW_D"))
        add_block("Bankruptcy / Winding Up", sccb.get("_SCCB_BKY_RAW"))
        ws3.sheet_state = "hidden"

    wb.save(out_path)

# ---------- CLI ----------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--stars", required=True)
    ap.add_argument("--cbs", required=True)
    ap.add_argument("--sccb", required=True)
    ap.add_argument("--out", default="Case_Output.xlsx")
    ap.add_argument("--adverse", action="store_true", help="If set, try adverse-news search using env keys.")
    args = ap.parse_args()

    for path in (args.stars, args.cbs, args.sccb):
        if not os.path.exists(path): sys.exit(f"File not found: {path}")

    texts = {
        "stars": read_pdf_text(args.stars),
        "cbs":   read_pdf_text(args.cbs),
        "sccb":  read_pdf_text(args.sccb)
    }

    stars = parse_stars(texts["stars"])
    cbs   = parse_cbs(texts["cbs"])
    sccb  = parse_sccb(texts["sccb"])

    # Optional adverse-news search
    hits = []
    if args.adverse:
        name_for_search = cbs.get("CBS Name") or sccb.get("Individual Name")
        if name_for_search:
            hits = adverse_news(name_for_search, limit=5)

    attachments = [os.path.basename(args.stars), os.path.basename(args.cbs), os.path.basename(args.sccb)]
    build_workbook(stars, sccb, cbs, args.out, attachments, adverse_hits=hits)
    print(f"Wrote {args.out} with sectioned sheets." + (f" Adverse-news: {len(hits)} item(s)." if hits else " Adverse-news: skipped or none."))

if __name__ == "__main__":
    main()
